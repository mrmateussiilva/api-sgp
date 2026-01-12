#!/usr/bin/env python3
"""
Script para gerar relatório financeiro diário e semanal em formato XLS.

Este script acessa o banco de dados SQLite, calcula valores totais por dia e semana,
e gera um relatório em Excel usando Polars.

Uso:
    python scripts/relatorio_financeiro.py [--data DATA] [--output OUTPUT.xlsx]
    
    --data: Data específica para relatório (formato: YYYY-MM-DD). Se não informado, usa hoje.
    --output: Caminho do arquivo de saída. Padrão: relatorio_financeiro_YYYY-MM-DD.xlsx
"""

import argparse
import sqlite3
import sys
import os
import json
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, Tuple

try:
    import polars as pl
except ImportError:
    print("❌ Erro: Polars não está instalado.")
    print("   Execute: uv pip install polars openpyxl")
    sys.exit(1)


def normalize_float_value(value: Optional[str]) -> float:
    """
    Normaliza valores de string para float.
    Lida com formatos como '370.00', '1,955.00', '1.955.00', '0.00', None, etc.
    """
    if value is None or value == '':
        return 0.0
    
    # Remove espaços e converte para string
    value_str = str(value).strip()
    
    if not value_str or value_str == 'None':
        return 0.0
    
    # Detecta formato e converte corretamente
    if ',' in value_str and '.' in value_str:
        # Verificar ordem: se vírgula vem depois do ponto, é formato brasileiro
        # Se ponto vem depois da vírgula, é formato misto
        pos_virgula = value_str.rfind(',')
        pos_ponto = value_str.rfind('.')
        if pos_ponto > pos_virgula:
            # Formato misto: 1,955.00 (vírgula como milhar, ponto como decimal)
            # Remove vírgula (separador de milhar) e mantém ponto (decimal)
            value_str = value_str.replace(',', '')
        else:
            # Formato brasileiro: 1.955,00 (ponto como milhar, vírgula como decimal)
            # Remove pontos de milhar e substitui vírgula por ponto decimal
            value_str = value_str.replace('.', '').replace(',', '.')
    elif ',' in value_str:
        # Formato brasileiro: 1.955,00 ou 1955,00
        # Remove pontos de milhar e substitui vírgula por ponto decimal
        value_str = value_str.replace('.', '').replace(',', '.')
    elif '.' in value_str:
        # Formato americano: pode ser 1955.00 ou 1.955.00 (com separador de milhar)
        parts = value_str.split('.')
        if len(parts) > 2:
            # Múltiplos pontos: o último é decimal, os anteriores são separadores de milhar
            # Ex: "1.955.00" -> parts = ["1", "955", "00"]
            # Remover pontos anteriores ao último, manter o último ponto
            decimal_part = parts[-1]
            integer_part = ''.join(parts[:-1])
            value_str = f"{integer_part}.{decimal_part}"
        elif len(parts) == 2:
            # Um ponto: verificar se é decimal (2-3 dígitos após) ou milhar
            if len(parts[1]) <= 3:
                # É decimal, mantém como está
                pass
            else:
                # Não é decimal, remove o ponto (é separador de milhar incorreto)
                value_str = value_str.replace('.', '')
        # Se len(parts) == 1, não há ponto (não deveria entrar aqui)
    
    try:
        return float(value_str)
    except (ValueError, TypeError):
        print(f"⚠️  Aviso: Valor não numérico encontrado: '{value}', usando 0.0")
        return 0.0


def get_db_path() -> Path:
    """
    Determina o caminho do banco de dados sempre usando shared.
    Prioriza API_ROOT/shared/db/banco.db se API_ROOT estiver definido,
    caso contrário usa diretório atual/shared/db/banco.db.
    """
    api_root = os.environ.get('API_ROOT')
    if api_root:
        db_path = Path(api_root) / 'shared' / 'db' / 'banco.db'
    else:
        # Usar diretório do script como base
        db_path = Path(__file__).parent.parent / 'shared' / 'db' / 'banco.db'
    
    if not db_path.exists():
        raise FileNotFoundError(
            f"❌ Banco de dados não encontrado em: {db_path}\n"
            f"   Certifique-se de que o diretório shared/db existe e contém banco.db"
        )
    
    return db_path


def get_week_start(date: datetime) -> datetime:
    """Retorna o início da semana (segunda-feira) para uma data."""
    days_since_monday = date.weekday()
    return date - timedelta(days=days_since_monday)


def fetch_pedidos_data(conn: sqlite3.Connection, data_inicio: Optional[str] = None, data_fim: Optional[str] = None) -> list:
    """
    Busca TODOS os dados de pedidos do banco de dados.
    
    Args:
        conn: Conexão SQLite
        data_inicio: Data de início (YYYY-MM-DD) ou None para todos
        data_fim: Data de fim (YYYY-MM-DD) ou None para todos
    
    Returns:
        Lista de tuplas com TODOS os campos da tabela pedidos na ordem:
        (id, numero, data_entrada, data_entrega, observacao, prioridade, status,
         cliente, telefone_cliente, cidade_cliente, valor_total, valor_frete, valor_itens,
         tipo_pagamento, obs_pagamento, forma_envio, forma_envio_id,
         financeiro, conferencia, sublimacao, costura, expedicao, pronto,
         sublimacao_maquina, sublimacao_data_impressao, items,
         data_criacao, ultima_atualizacao)
    """
    query = """
        SELECT 
            id,
            numero,
            data_entrada,
            data_entrega,
            observacao,
            prioridade,
            COALESCE(status, 'pendente') as status,
            cliente,
            telefone_cliente,
            cidade_cliente,
            COALESCE(valor_total, '0.00') as valor_total,
            COALESCE(valor_frete, '0.00') as valor_frete,
            COALESCE(valor_itens, '0.00') as valor_itens,
            tipo_pagamento,
            obs_pagamento,
            forma_envio,
            forma_envio_id,
            financeiro,
            conferencia,
            sublimacao,
            costura,
            expedicao,
            pronto,
            sublimacao_maquina,
            sublimacao_data_impressao,
            items,
            data_criacao,
            ultima_atualizacao
        FROM pedidos
        WHERE 1=1
    """
    
    params = []
    if data_inicio and data_fim:
        if data_inicio == data_fim:
            # Se for a mesma data, usar igualdade direta
            query += " AND data_entrada = ?"
            params.append(data_inicio)
        else:
            # Se for um período, usar >= e <=
            query += " AND data_entrada >= ? AND data_entrada <= ?"
            params.append(data_inicio)
            params.append(data_fim)
    elif data_inicio:
        query += " AND data_entrada >= ?"
        params.append(data_inicio)
    elif data_fim:
        query += " AND data_entrada <= ?"
        params.append(data_fim)
    
    query += " ORDER BY data_entrada, id"
    
    cursor = conn.execute(query, params)
    result = cursor.fetchall()
    
    # Debug: verificar se não encontrou resultados
    if not result and (data_inicio or data_fim):
        print(f"   🔍 Query executada: {query}")
        print(f"   🔍 Parâmetros: {params}")
        # Verificar se há pedidos no banco
        cursor2 = conn.execute("SELECT COUNT(*) FROM pedidos")
        total = cursor2.fetchone()[0]
        print(f"   🔍 Total de pedidos no banco: {total}")
        if total > 0:
            # Mostrar algumas datas disponíveis
            cursor3 = conn.execute("SELECT DISTINCT data_entrada FROM pedidos ORDER BY data_entrada DESC LIMIT 10")
            datas = cursor3.fetchall()
            print(f"   🔍 Datas disponíveis (últimas 10): {[d[0] for d in datas]}")
            # Verificar se há pedidos na data específica
            if data_inicio == data_fim:
                cursor4 = conn.execute("SELECT COUNT(*) FROM pedidos WHERE data_entrada = ?", (data_inicio,))
                count_exato = cursor4.fetchone()[0]
                print(f"   🔍 Pedidos com data_entrada = '{data_inicio}': {count_exato}")
    
    return result


def create_pedidos_detalhados(pedidos_data: list) -> pl.DataFrame:
    """
    Cria DataFrame com TODOS os detalhes de cada pedido (uma linha por pedido).
    Mapeia TODOS os campos da tabela pedidos para análise completa.
    """
    data = []
    
    for row in pedidos_data:
        # Desempacotar todos os campos na ordem da query
        (pedido_id, numero, data_entrada, data_entrega, observacao, prioridade, status,
         cliente, telefone_cliente, cidade_cliente, valor_total, valor_frete, valor_itens,
         tipo_pagamento, obs_pagamento, forma_envio, forma_envio_id,
         financeiro, conferencia, sublimacao, costura, expedicao, pronto,
         sublimacao_maquina, sublimacao_data_impressao, items_json,
         data_criacao, ultima_atualizacao) = row
        
        # Normalizar valores monetários
        total = normalize_float_value(valor_total)
        frete = normalize_float_value(valor_frete)
        itens = normalize_float_value(valor_itens)
        
        # Se valor_itens não estiver preenchido, calcular como total - frete
        if itens == 0.0 and total > 0:
            itens = total - frete
        
        total_calculado = frete + itens
        diferenca = total - total_calculado
        
        # Decodificar cidade_cliente se tiver separador de estado
        cidade_final = cidade_cliente or ''
        estado_cliente = ''
        if cidade_cliente and '||' in cidade_cliente:
            parts = cidade_cliente.split('||', 1)
            cidade_final = parts[0].strip()
            estado_cliente = parts[1].strip() if len(parts) > 1 else ''
        
        # Converter booleanos para texto legível
        financeiro_txt = 'Sim' if financeiro else 'Não'
        conferencia_txt = 'Sim' if conferencia else 'Não'
        sublimacao_txt = 'Sim' if sublimacao else 'Não'
        costura_txt = 'Sim' if costura else 'Não'
        expedicao_txt = 'Sim' if expedicao else 'Não'
        pronto_txt = 'Sim' if pronto else 'Não'
        
        # Contar itens do JSON
        qtd_itens = 0
        try:
            if items_json:
                items_list = json.loads(items_json) if isinstance(items_json, str) else items_json
                qtd_itens = len(items_list) if isinstance(items_list, list) else 0
        except:
            qtd_itens = 0
        
        # Formatar datas para exibição
        data_criacao_str = ''
        if data_criacao:
            try:
                if isinstance(data_criacao, str):
                    dt = datetime.fromisoformat(data_criacao.replace('Z', '+00:00'))
                else:
                    dt = data_criacao
                data_criacao_str = dt.strftime('%Y-%m-%d %H:%M:%S')
            except:
                data_criacao_str = str(data_criacao)
        
        ultima_atualizacao_str = ''
        if ultima_atualizacao:
            try:
                if isinstance(ultima_atualizacao, str):
                    dt = datetime.fromisoformat(ultima_atualizacao.replace('Z', '+00:00'))
                else:
                    dt = ultima_atualizacao
                ultima_atualizacao_str = dt.strftime('%Y-%m-%d %H:%M:%S')
            except:
                ultima_atualizacao_str = str(ultima_atualizacao)
        
        data.append({
            # Identificação
            'ID': pedido_id,
            'Número': numero or '',
            
            # Datas
            'Data Entrada': data_entrada or '',
            'Data Entrega': data_entrega or '',
            'Data Criação': data_criacao_str,
            'Última Atualização': ultima_atualizacao_str,
            
            # Cliente
            'Cliente': cliente or '',
            'Telefone Cliente': telefone_cliente or '',
            'Cidade Cliente': cidade_final,
            'Estado Cliente': estado_cliente,
            
            # Status e Prioridade
            'Status': status or '',
            'Prioridade': prioridade or 'NORMAL',
            'Observação': observacao or '',
            
            # Valores Financeiros
            'Valor Total': total,
            'Frete': frete,
            'Serviços': itens,
            'Total Calculado': total_calculado,
            'Diferença': diferenca,  # Deve ser 0 se estiver correto
            
            # Pagamento
            'Tipo Pagamento': tipo_pagamento or '',
            'Obs Pagamento': obs_pagamento or '',
            
            # Envio
            'Forma Envio': forma_envio or '',
            'Forma Envio ID': forma_envio_id,
            
            # Status de Produção (Booleanos)
            'Financeiro': financeiro_txt,
            'Conferência': conferencia_txt,
            'Sublimação': sublimacao_txt,
            'Costura': costura_txt,
            'Expedição': expedicao_txt,
            'Pronto': pronto_txt,
            
            # Sublimação Detalhes
            'Sublimação Máquina': sublimacao_maquina or '',
            'Sublimação Data Impressão': sublimacao_data_impressao or '',
            
            # Items
            'Qtd Items': qtd_itens,
            'Items (JSON)': 'Sim' if items_json else 'Não'
        })
    
    if not data:
        return pl.DataFrame({
            'ID': [], 'Número': [], 'Data Entrada': [], 'Data Entrega': [],
            'Cliente': [], 'Status': [], 'Valor Total': [], 'Frete': [],
            'Serviços': [], 'Total Calculado': [], 'Diferença': []
        })
    
    return pl.DataFrame(data).sort('ID')


def calculate_daily_totals(pedidos_data: list) -> pl.DataFrame:
    """Calcula totais por dia."""
    data = []
    
    for row in pedidos_data:
        # Desempacotar campos necessários (ajustar índices para nova estrutura)
        pedido_id, numero, data_entrada, data_entrega, observacao, prioridade, status, \
        cliente, telefone_cliente, cidade_cliente, valor_total, valor_frete, valor_itens, \
        tipo_pagamento, obs_pagamento, forma_envio, forma_envio_id, \
        financeiro, conferencia, sublimacao, costura, expedicao, pronto, \
        sublimacao_maquina, sublimacao_data_impressao, items_json, \
        data_criacao, ultima_atualizacao = row
        
        # Normalizar valores
        total = normalize_float_value(valor_total)
        frete = normalize_float_value(valor_frete)
        itens = normalize_float_value(valor_itens)
        
        # Se valor_itens não estiver preenchido, calcular como total - frete
        if itens == 0.0 and total > 0:
            itens = total - frete
        
        data.append({
            'data': data_entrada,
            'valor_total': total,
            'valor_frete': frete,
            'valor_itens': itens,
            'qtd_pedidos': 1
        })
    
    if not data:
        return pl.DataFrame({
            'data': [],
            'valor_total': [],
            'valor_frete': [],
            'valor_itens': [],
            'qtd_pedidos': []
        })
    
    df = pl.DataFrame(data)
    
    # Agrupar por data
    daily_totals = df.group_by('data').agg([
        pl.sum('valor_total').alias('total_dia'),
        pl.sum('valor_frete').alias('frete_dia'),
        pl.sum('valor_itens').alias('servicos_dia'),
        pl.sum('qtd_pedidos').alias('qtd_pedidos')
    ]).sort('data')
    
    # Adicionar coluna de total calculado (frete + serviços)
    daily_totals = daily_totals.with_columns([
        (pl.col('frete_dia') + pl.col('servicos_dia')).alias('total_calculado')
    ])
    
    return daily_totals


def calculate_weekly_totals(pedidos_data: list) -> pl.DataFrame:
    """Calcula totais por semana (segunda a domingo)."""
    data = []
    
    for row in pedidos_data:
        # Desempacotar campos necessários (ajustar índices para nova estrutura)
        pedido_id, numero, data_entrada, data_entrega, observacao, prioridade, status, \
        cliente, telefone_cliente, cidade_cliente, valor_total, valor_frete, valor_itens, \
        tipo_pagamento, obs_pagamento, forma_envio, forma_envio_id, \
        financeiro, conferencia, sublimacao, costura, expedicao, pronto, \
        sublimacao_maquina, sublimacao_data_impressao, items_json, \
        data_criacao, ultima_atualizacao = row
        
        # Normalizar valores
        total = normalize_float_value(valor_total)
        frete = normalize_float_value(valor_frete)
        itens = normalize_float_value(valor_itens)
        
        # Se valor_itens não estiver preenchido, calcular como total - frete
        if itens == 0.0 and total > 0:
            itens = total - frete
        
        # Converter data_entrada para datetime
        semana_str = "Data inválida"
        semana_inicio = ''
        
        try:
            date_obj = datetime.strptime(data_entrada, '%Y-%m-%d').date()
            week_start = get_week_start(date_obj)
            week_end = week_start + timedelta(days=6)
            semana_str = f"{week_start.strftime('%Y-%m-%d')} a {week_end.strftime('%Y-%m-%d')}"
            semana_inicio = week_start.strftime('%Y-%m-%d')
        except (ValueError, TypeError) as e:
            print(f"⚠️  Aviso: Erro ao processar data '{data_entrada}': {e}")
        
        data.append({
            'semana': semana_str,
            'semana_inicio': semana_inicio,
            'data': data_entrada,
            'valor_total': total,
            'valor_frete': frete,
            'valor_itens': itens,
            'qtd_pedidos': 1
        })
    
    if not data:
        return pl.DataFrame({
            'semana': [],
            'semana_inicio': [],
            'total_semana': [],
            'frete_semana': [],
            'servicos_semana': [],
            'qtd_pedidos': []
        })
    
    df = pl.DataFrame(data)
    
    # Agrupar por semana
    weekly_totals = df.group_by('semana', 'semana_inicio').agg([
        pl.sum('valor_total').alias('total_semana'),
        pl.sum('valor_frete').alias('frete_semana'),
        pl.sum('valor_itens').alias('servicos_semana'),
        pl.sum('qtd_pedidos').alias('qtd_pedidos')
    ]).sort('semana_inicio')
    
    # Adicionar coluna de total calculado (frete + serviços)
    weekly_totals = weekly_totals.with_columns([
        (pl.col('frete_semana') + pl.col('servicos_semana')).alias('total_calculado')
    ])
    
    # Remover coluna semana_inicio do resultado final (usada apenas para ordenação)
    weekly_totals = weekly_totals.select([
        'semana',
        'total_semana',
        'frete_semana',
        'servicos_semana',
        'total_calculado',
        'qtd_pedidos'
    ])
    
    return weekly_totals


def format_currency(value: float) -> str:
    """Formata valor como moeda brasileira (R$ 1.234,56)."""
    if value is None:
        value = 0.0
    
    # Formatação manual para evitar dependência de locale
    valor_str = f"{abs(value):,.2f}"
    
    # Separar parte inteira e decimal
    parts = valor_str.split('.')
    parte_inteira = parts[0].replace(',', '.')  # Pontos de milhar
    parte_decimal = parts[1] if len(parts) > 1 else '00'
    
    # Montar no formato brasileiro
    sinal = '-' if value < 0 else ''
    return f"{sinal}R$ {parte_inteira},{parte_decimal}"


def generate_report(output_path: Path, data_inicio: Optional[str] = None, data_fim: Optional[str] = None) -> None:
    """
    Gera relatório financeiro em XLS.
    
    Args:
        output_path: Caminho do arquivo de saída
        data_inicio: Data de início para filtro (YYYY-MM-DD) ou None
        data_fim: Data de fim para filtro (YYYY-MM-DD) ou None
    """
    print("=" * 70)
    print("📊 Geração de Relatório Financeiro")
    print("=" * 70)
    
    # Obter caminho do banco
    try:
        db_path = get_db_path()
        print(f"📁 Banco de dados: {db_path}")
    except FileNotFoundError as e:
        print(str(e))
        sys.exit(1)
    
    # Conectar ao banco
    try:
        conn = sqlite3.connect(str(db_path))
        print("✅ Conectado ao banco de dados")
    except sqlite3.Error as e:
        print(f"❌ Erro ao conectar ao banco: {e}")
        sys.exit(1)
    
    # Buscar dados
    print(f"\n📥 Buscando dados de pedidos...")
    if data_inicio:
        print(f"   Data início: {data_inicio}")
    if data_fim:
        print(f"   Data fim: {data_fim}")
    
    try:
        pedidos_data = fetch_pedidos_data(conn, data_inicio, data_fim)
        print(f"✅ {len(pedidos_data)} pedidos encontrados")
    except sqlite3.Error as e:
        print(f"❌ Erro ao buscar dados: {e}")
        conn.close()
        sys.exit(1)
    finally:
        conn.close()
    
    if not pedidos_data:
        print("⚠️  Nenhum pedido encontrado no período especificado.")
        sys.exit(0)
    
    # Criar planilha detalhada (uma linha por pedido)
    print("\n📊 Criando planilha detalhada de pedidos...")
    pedidos_detalhados = create_pedidos_detalhados(pedidos_data)
    
    # Calcular totais
    print("📊 Calculando totais por dia...")
    daily_totals = calculate_daily_totals(pedidos_data)
    
    print("📊 Calculando totais por semana...")
    weekly_totals = calculate_weekly_totals(pedidos_data)
    
    # Criar resumo geral (ajustar índices: agora row[10] é valor_total, row[11] é valor_frete, row[12] é valor_itens)
    total_geral = sum(normalize_float_value(row[10]) for row in pedidos_data)
    total_frete = sum(normalize_float_value(row[11]) for row in pedidos_data)
    total_servicos = sum(normalize_float_value(row[12]) for row in pedidos_data)
    
    # Se valor_itens não estiver preenchido, calcular
    if total_servicos == 0 and total_geral > 0:
        total_servicos = total_geral - total_frete
    
    resumo_geral = pl.DataFrame({
        'descricao': ['Total Geral', 'Frete Total', 'Serviços Total', 'Total Calculado'],
        'valor': [total_geral, total_frete, total_servicos, total_frete + total_servicos]
    })
    
    # Preparar dados para exportação (manter valores numéricos para cálculos no Excel)
    # Mas também criar versões formatadas para exibição
    daily_formatted = daily_totals.with_columns([
        pl.col('total_dia').alias('Total (R$)'),
        pl.col('frete_dia').alias('Frete (R$)'),
        pl.col('servicos_dia').alias('Serviços (R$)'),
        pl.col('total_calculado').alias('Total Calculado (R$)'),
        pl.col('qtd_pedidos').alias('Quantidade Pedidos')
    ]).select([
        'data',
        'Total (R$)',
        'Frete (R$)',
        'Serviços (R$)',
        'Total Calculado (R$)',
        'Quantidade Pedidos'
    ])
    
    weekly_formatted = weekly_totals.with_columns([
        pl.col('total_semana').alias('Total (R$)'),
        pl.col('frete_semana').alias('Frete (R$)'),
        pl.col('servicos_semana').alias('Serviços (R$)'),
        pl.col('total_calculado').alias('Total Calculado (R$)'),
        pl.col('qtd_pedidos').alias('Quantidade Pedidos')
    ]).select([
        'semana',
        'Total (R$)',
        'Frete (R$)',
        'Serviços (R$)',
        'Total Calculado (R$)',
        'Quantidade Pedidos'
    ])
    
    resumo_formatted = resumo_geral.with_columns([
        pl.col('valor').alias('Valor (R$)')
    ]).select(['descricao', 'Valor (R$)'])
    
    # Exportar para Excel
    print(f"\n💾 Gerando arquivo Excel: {output_path}")
    
    try:
        # Verificar se openpyxl está disponível (necessário para múltiplas planilhas)
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            print(f"❌ Erro: openpyxl não está instalado.")
            print(f"   Execute: uv pip install openpyxl")
            raise
        
        # Converter DataFrames do Polars para dicionários/listas para escrita no Excel
        # Criar workbook com openpyxl
        wb = Workbook()
        wb.remove(wb.active)  # Remover planilha padrão
        
        # Planilha 1: Resumo Geral
        ws_resumo = wb.create_sheet("Resumo Geral")
        # Cabeçalhos
        ws_resumo.append(['Descrição', 'Valor (R$)'])
        # Dados (usar valores numéricos do resumo_geral)
        resumo_dict = resumo_geral.to_dict(as_series=False)
        for i in range(len(resumo_dict['descricao'])):
            ws_resumo.append([
                resumo_dict['descricao'][i],
                resumo_dict['valor'][i]  # Valor numérico (Excel formatará)
            ])
        
        # Planilha 2: Totais por Dia
        ws_dia = wb.create_sheet("Totais por Dia")
        # Cabeçalhos
        ws_dia.append(['Data', 'Total (R$)', 'Frete (R$)', 'Serviços (R$)', 'Total Calculado (R$)', 'Quantidade Pedidos'])
        # Dados (usar valores numéricos do daily_totals)
        daily_dict = daily_totals.to_dict(as_series=False)
        for i in range(len(daily_dict['data'])):
            ws_dia.append([
                daily_dict['data'][i],
                daily_dict['total_dia'][i],
                daily_dict['frete_dia'][i],
                daily_dict['servicos_dia'][i],
                daily_dict['total_calculado'][i],
                daily_dict['qtd_pedidos'][i]
            ])
        
        # Planilha 3: Totais por Semana
        ws_semana = wb.create_sheet("Totais por Semana")
        # Cabeçalhos
        ws_semana.append(['Semana', 'Total (R$)', 'Frete (R$)', 'Serviços (R$)', 'Total Calculado (R$)', 'Quantidade Pedidos'])
        # Dados (usar valores numéricos do weekly_totals)
        weekly_dict = weekly_totals.to_dict(as_series=False)
        for i in range(len(weekly_dict['semana'])):
            ws_semana.append([
                weekly_dict['semana'][i],
                weekly_dict['total_semana'][i],
                weekly_dict['frete_semana'][i],
                weekly_dict['servicos_semana'][i],
                weekly_dict['total_calculado'][i],
                weekly_dict['qtd_pedidos'][i]
            ])
        
        # Planilha 4: Detalhes dos Pedidos (uma linha por pedido com TODOS os campos)
        ws_detalhes = wb.create_sheet("Detalhes dos Pedidos")
        # Cabeçalhos - usar todas as colunas do DataFrame
        pedidos_dict = pedidos_detalhados.to_dict(as_series=False)
        colunas = list(pedidos_dict.keys())
        ws_detalhes.append(colunas)
        
        # Dados (uma linha por pedido com TODOS os campos)
        num_rows = len(pedidos_dict['ID'])
        for i in range(num_rows):
            linha = [pedidos_dict[col][i] for col in colunas]
            ws_detalhes.append(linha)
        
        # Ajustar largura das colunas para melhor visualização
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(colunas, start=1):
            col_letter = get_column_letter(idx)
            # Ajustar largura baseada no conteúdo
            max_length = len(str(col))  # Tamanho mínimo = tamanho do cabeçalho
            for row_idx in range(2, min(ws_detalhes.max_row + 1, 102)):  # Verificar até 100 linhas
                cell_value = ws_detalhes[f"{col_letter}{row_idx}"].value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            ws_detalhes.column_dimensions[col_letter].width = min(max_length + 2, 50)  # Máximo 50 caracteres
        
        # Salvar arquivo
        wb.save(str(output_path))
        
        print(f"✅ Relatório gerado com sucesso!")
        print(f"   📄 Arquivo: {output_path}")
        
        # Mostrar resumo no console
        print("\n" + "=" * 70)
        print("📊 RESUMO GERAL")
        print("=" * 70)
        print(f"   Total Geral:        {format_currency(total_geral)}")
        print(f"   Frete Total:        {format_currency(total_frete)}")
        print(f"   Serviços Total:     {format_currency(total_servicos)}")
        print(f"   Total Calculado:    {format_currency(total_frete + total_servicos)}")
        print(f"   Quantidade Pedidos: {len(pedidos_data)}")
        
        if len(daily_totals) > 0:
            print(f"\n   Período: {daily_totals['data'].min()} a {daily_totals['data'].max()}")
        
        print("=" * 70)
        
    except Exception as e:
        print(f"❌ Erro ao gerar arquivo Excel: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description="Gera relatório financeiro diário e semanal em formato XLS",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemplos:
  # Relatório de hoje
  python scripts/relatorio_financeiro.py
  
  # Relatório de uma data específica
  python scripts/relatorio_financeiro.py --data 2026-01-10
  
  # Relatório de um período
  python scripts/relatorio_financeiro.py --data-inicio 2026-01-01 --data-fim 2026-01-31
  
  # Especificar arquivo de saída
  python scripts/relatorio_financeiro.py --output relatorio_janeiro.xlsx
        """
    )
    
    parser.add_argument(
        '--data',
        type=str,
        help='Data específica para relatório (formato: YYYY-MM-DD). Gera relatório apenas desta data.'
    )
    
    parser.add_argument(
        '--data-inicio',
        type=str,
        help='Data de início do período (formato: YYYY-MM-DD)'
    )
    
    parser.add_argument(
        '--data-fim',
        type=str,
        help='Data de fim do período (formato: YYYY-MM-DD)'
    )
    
    parser.add_argument(
        '--output',
        type=str,
        help='Caminho do arquivo de saída (padrão: relatorio_financeiro_YYYY-MM-DD.xlsx)'
    )
    
    args = parser.parse_args()
    
    # Determinar período
    data_inicio = None
    data_fim = None
    
    if args.data:
        # Data específica
        data_inicio = args.data
        data_fim = args.data
    else:
        # Usar argumentos de período ou padrão (hoje)
        if args.data_inicio:
            data_inicio = args.data_inicio
        if args.data_fim:
            data_fim = args.data_fim
        
        # Se não especificado, usar hoje como padrão
        if not data_inicio and not data_fim:
            hoje = datetime.now().date().isoformat()
            data_inicio = hoje
            data_fim = hoje
    
    # Determinar arquivo de saída
    if args.output:
        output_path = Path(args.output)
    else:
        data_str = data_inicio if data_inicio else datetime.now().date().isoformat()
        output_path = Path(__file__).parent.parent / 'relatorios' / f'relatorio_financeiro_{data_str}.xlsx'
    
    # Criar diretório se não existir
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Gerar relatório
    generate_report(output_path, data_inicio, data_fim)


if __name__ == "__main__":
    import os
    main()

